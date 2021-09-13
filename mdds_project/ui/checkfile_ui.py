# -*- coding: utf-8 -*-
from __future__ import division
import env
import sqlite3
from PyQt4 import QtGui,QtCore
from PyQt4.QtGui import *
from PyQt4.QtCore import *
import sys
import os
import threading
from checkfile2 import Ui_Form2
from checkfile import Ui_Form
import mdds.multi_thread_detect
from mdds.lib.configMgr import config
import mdds.detect
from mdds.lib.dbMgr import DBMgr
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

import thread

QTextCodec.setCodecForTr(QTextCodec.codecForName("utf8"))
#DATABASE = config.DATABASE
#DBManager = DBMgr()
DBLock = threading.Lock()
count = 0
max_num = 0

try:
    _fromUtf8 = QtCore.QString.fromUtf8
except AttributeError:
    def _fromUtf8(s):
        return s

class MyWindow(QtGui.QMainWindow):
    @pyqtSlot(int)
    def updateLabel(self,value):
        progresstext = str(value) + "/" + str(max_num)
        self.ui.progressLable.setText(progresstext)

    def __init__(self):
        super(MyWindow,self).__init__()
        self.filename = ''
        self.ui = Ui_Form()
        self.source = ''
        self.ui.setupUi(self)
        self.model = QtGui.QStandardItemModel(self.ui.tableView)
        self.model.setRowCount(count)
        self.model.setColumnCount(4)

        self.model.setHeaderData(0, QtCore.Qt.Horizontal, _fromUtf8(u"文件名"))
        self.model.setHeaderData(1, QtCore.Qt.Horizontal, _fromUtf8(u"文件类型"))
        self.model.setHeaderData(2, QtCore.Qt.Horizontal, _fromUtf8(u"是否为恶意文档"))
        self.model.setHeaderData(3, QtCore.Qt.Horizontal, _fromUtf8(u"描述"))
        self.ui.tableView.setModel(self.model)
        self.ui.tableView.horizontalHeader().setStretchLastSection(True)
        self.ui.tableView.setUpdatesEnabled(True)
        self.ui.tableView.setSortingEnabled(True)
        self.ui.tableView.setEditTriggers(QtGui.QAbstractItemView.NoEditTriggers)
        self.connect(self.ui.fileButton,QtCore.SIGNAL('clicked()'),self.showFileName)
        self.connect(self.ui.directoryButton, QtCore.SIGNAL('clicked()'), self.showDirectoryName)
        self.connect(self.ui.startButton,QtCore.SIGNAL('clicked()'),self.startCheckFile)
        self.connect(self.ui.resultButton, QtCore.SIGNAL('clicked()'), self.save_record)


    def showFileName(self):
        self.filename = QtGui.QFileDialog.getOpenFileName(self,self.tr('选择文件'),'./')
        self.source = unicode(self.filename.toUtf8(), 'utf-8', 'ignore')
        self.ui.lineEdit.setText(self.filename)

    def showDirectoryName(self):
        self.filename = QtGui.QFileDialog.getExistingDirectory(self,self.tr('选择文件夹'),'./')
        self.source = unicode(self.filename.toUtf8(), 'utf-8', 'ignore')
        self.ui.lineEdit.setText(self.filename)

    def outSelect(self):
        row = self.ui.tableView.currentIndex().row()
        col = self.ui.tableView.currentIndex().column()
        if col == 3 and str(self.model.item(row, 2).text()) == 'True':
            UIManager2.ui2.textEdit_3.setText(str(self.model.item(row, col).text()))
            UIManager2.ui2.lineEdit_2.setText(str(self.model.item(row, 1).text()))
            UIManager2.ui2.lineEdit.setText(str(self.model.item(row, 0).text()))
            UIManager2.showEdit()
        if col == 0 and str(self.model.item(row, 2).text()) == 'True':
            UIManager2.ui2.textEdit_3.setText(str(self.model.item(row, 3).text()))
            UIManager2.ui2.lineEdit_2.setText(str(self.model.item(row, 1).text()))
            UIManager2.ui2.lineEdit.setText(str(self.model.item(row, col).text()))
            UIManager2.showEdit()
        else:
            return

    def save_record(self):
        reply = QMessageBox.question(self, self.tr('提示'), self.tr('是否保存检查结果到本地？'),
                                     QMessageBox.Yes, QMessageBox.No)
        if reply == QMessageBox.Yes:
            if self.model.horizontalHeaderItem(1):
                wbk = Workbook()
                ws = wbk.active
                _ = ws.cell(row=1, column=1,
                            value=unicode(self.model.horizontalHeaderItem(0).text().toUtf8(), 'utf-8', 'ignore'))
                _ = ws.cell(row=1, column=2,
                            value=unicode(self.model.horizontalHeaderItem(1).text().toUtf8(), 'utf-8', 'ignore'))
                _ = ws.cell(row=1, column=3,
                            value=unicode(self.model.horizontalHeaderItem(2).text().toUtf8(), 'utf-8', 'ignore'))
                _ = ws.cell(row=1, column=4,
                            value=unicode(self.model.horizontalHeaderItem(3).text().toUtf8(), 'utf-8', 'ignore'))
                filename1 = QtGui.QFileDialog.getSaveFileName(self, self.tr('文件保存'), './test.xlsx',
                                                          "Excel files(*.xls*.xlsx)")
                filename1 = unicode(filename1.toUtf8(), 'utf-8', 'ignore')
                if ws:
                    row = self.model.rowCount()
                    column = self.model.columnCount()
                    for i in range(2, row + 2):
                        for j in range(1, column + 1):
                            _ = ws.cell(row=i, column=j,
                                        value=unicode(self.model.item(i - 2, j - 1).text().toUtf8(), 'utf-8', 'ignore'))
                    try:
                        wbk.save(filename1)

                        return QMessageBox.warning(self, self.tr('提示'), self.tr('存储成功'))

                    except:
                        pass

            else:
                return QMessageBox.warning(self,self.tr('提示'), self.tr('请先检测文件'))
        else:
            pass

    def visitDir(self,path):
        global max_num
        if os.path.isfile(path):
            max_num = 1
            return
        for dir, folder, file in os.walk(path):
            for i in file:
                max_num = max_num + 1




    def startCheckFile(self):

        global max_num
        if not self.ui.lineEdit.text():
            return QMessageBox.warning(self,self.tr('提示'),self.tr('请选择文件或文件夹'))

        if self.model.rowCount() >= 1:
            self.ui.progressBar.setValue(0)
            self.model.removeRows(0, self.model.rowCount())
        else:
            self.visitDir(self.source)
            # t= threading.Thread(target = mdds.multi_thread_detect.main,args = (self.source, myCalbFunc))
            # t.start()
            # t.join()
            self.analyseThread = AnalyzeThread(self.source)
            self.connect(self.analyseThread, SIGNAL("finished()"),self.done)
            # self.analyseThread._signal.connect(self.callback)
            self.analyseThread.start()
            #print(max_num)
            # self.clear()
            # result = QMessageBox.warning(self, self.tr('提示'), self.tr('检查完毕'))
            # if result:
            #     global count
            #     # count = 0
            #     # mdds.detect.sum = 0
            #     # max_num = 0
            self.connect(self.ui.tableView, SIGNAL('clicked (QModelIndex)'), self.outSelect)

    def callback(self):
        progress_text = str(mdds.detect.sum) + "/" + str(max_num)
        UIManager.ui.progressLable.setText(progress_text)

    def done(self):
        self.clear()
        result = QMessageBox.warning(self, self.tr('提示'), self.tr('检查完毕'))
        if result:
            global count, max_num
            count = 0
            mdds.detect.sum = 0
            max_num = 0

    def clear(self):
        self.ui.lineEdit.setText('')
        self.source = ''
        self.filename = ''


    '''
    def load_all(self):
        conn = sqlite3.connect(":memory:", check_same_thread=False)
        cur = conn.cursor()

        rows = cur.fetchall()
        row = len(rows)  # 取得记录个数，用于设置表格的行数
        vol = len(rows[0])  # 取得字段数，用于设置表格的列数
        cur.close()
        conn.close()
        self.connect(self.ui.tableView, SIGNAL('clicked (QModelIndex)'), self.outSelect)
        self.model.setRowCount(row)
        self.model.setColumnCount(vol)
        self.model.setHeaderData(0, QtCore.Qt.Horizontal, _fromUtf8(u"文件名"))
        self.model.setHeaderData(1, QtCore.Qt.Horizontal, _fromUtf8(u"文件类型"))
        self.model.setHeaderData(2, QtCore.Qt.Horizontal, _fromUtf8(u"是否为恶意文档"))
        self.model.setHeaderData(3, QtCore.Qt.Horizontal, _fromUtf8(u"描述"))
        self.ui.tableView.setModel(self.model)
        self.ui.tableView.horizontalHeader().setStretchLastSection(True)
        for i in range(row):
            for j in range(vol):
                temp_data = rows[i][j]  # 临时记录，不能直接插入表格
                data = QStandardItem(str(temp_data))  # 转换后可插入表格
                self.model.setItem(i, j, data)
        self.clear()
        return QMessageBox.warning(self, self.tr('提示'), self.tr('显示完毕'))
    '''

    def showRecord(self,result):
        global count
        self.model.setItem(count, 0, QtGui.QStandardItem(result[0]))
        self.model.setItem(count, 1, QtGui.QStandardItem(result[1]))
        self.model.setItem(count, 2, QtGui.QStandardItem(result[2]))
        self.model.setItem(count, 3, QtGui.QStandardItem(result[3]))
        self.model.item(count, 0).setTextAlignment(QtCore.Qt.AlignCenter)
        self.model.item(count, 1).setTextAlignment(QtCore.Qt.AlignCenter)
        self.model.item(count, 2).setTextAlignment(QtCore.Qt.AlignCenter)
        self.model.item(count, 3).setTextAlignment(QtCore.Qt.AlignCenter)
        if result[2] == 'True' and result[1] == 'DOCX' and result[3] == '':
            self.model.setItem(count, 3, QtGui.QStandardItem('May be have a bad structure'))
        if result[2] == 'True' and result[1] == 'PDF' and result[3] == '':
            self.model.setItem(count, 3, QtGui.QStandardItem('May be have a bad structure'))
        if result[2] == 'True':
            self.model.item(count, 0).setForeground(QtGui.QBrush(QtGui.QColor(255, 0, 0)))
            self.model.item(count, 1).setForeground(QtGui.QBrush(QtGui.QColor(255, 0, 0)))
            self.model.item(count, 2).setForeground(QtGui.QBrush(QtGui.QColor(255, 0, 0)))
            self.model.item(count, 3).setForeground(QtGui.QBrush(QtGui.QColor(255, 0, 0)))


        count = count + 1


    def closeEvent(self, *args, **kwargs):
        reply = QMessageBox.question(self, self.tr('提示'), self.tr('是否保存检查结果到本地？'),
                                     QMessageBox.Yes, QMessageBox.No)
        if reply == QMessageBox.Yes:
            if self.model.horizontalHeaderItem(1):
                wbk = Workbook()
                ws = wbk.active
                _ = ws.cell(row=1, column=1,
                            value=unicode(self.model.horizontalHeaderItem(0).text().toUtf8(), 'utf-8', 'ignore'))
                _ = ws.cell(row=1, column=2,
                            value=unicode(self.model.horizontalHeaderItem(1).text().toUtf8(), 'utf-8', 'ignore'))
                _ = ws.cell(row=1, column=3,
                            value=unicode(self.model.horizontalHeaderItem(2).text().toUtf8(), 'utf-8', 'ignore'))
                _ = ws.cell(row=1, column=4,
                            value=unicode(self.model.horizontalHeaderItem(3).text().toUtf8(), 'utf-8', 'ignore'))
                filename1 = QtGui.QFileDialog.getSaveFileName(self, self.tr('文件保存'), './test.xlsx',
                                                          "Excel files(*.xls*.xlsx)")
                filename1 = unicode(filename1.toUtf8(), 'utf-8', 'ignore')
                if ws:
                    row = self.model.rowCount()
                    column = self.model.columnCount()
                    for i in range(2, row + 2):
                        for j in range(1, column + 1):
                            _ = ws.cell(row=i, column=j,
                                        value=unicode(self.model.item(i - 2, j - 1).text().toUtf8(), 'utf-8', 'ignore'))
                    try:
                        wbk.save(filename1)
                        return QMessageBox.warning(self, self.tr('提示'), self.tr('存储成功'))
                    except:
                        pass

            else:
                return QMessageBox.warning(self,self.tr('提示'), self.tr('无可存储文件'))
        else:
            pass

    # def myCalbFunc(self, result):
    #     """对检测结果进行处理的函数，可以是界面展示的回调函数，存储数据库的回调函数等
    #         本函数是将检测结果保存到了数据库中
    #         输入：
    #             result: [filename, fileType, BAD/GOOD, describe]
    #     """
    #     global max_num
    #     DBLock.acquire()
    #     # # print(mdds.detect.sum)
    #     # print(max_num)
    #     progress = (mdds.detect.sum / max_num) * 100
    #     progress_text = str(mdds.detect.sum) + "/" + str(max_num)
    #     # UIManager.ui.progressLineEdit.setText(progress_text)
    #     UIManager.ui.progressBar.setValue(progress)
    #     UIManager.ui.tableView.setModel(UIManager.model)
    #     result[2] = str(result[2])
    #     UIManager.showRecord(result)
    #     DBLock.release()

class MyWindow2(QtGui.QMainWindow):

    def __init__(self):
        super(MyWindow2,self).__init__()
        self.ui2 = Ui_Form2()
        self.ui2.setupUi(self)

    def showEdit(self):
        if not self.isVisible():
            UIManager2.show()


class AnalyzeThread(QtCore.QThread):
    _signal = pyqtSignal()
    def __init__(self, source,parent=None):
        QThread.__init__(self)
        self.source = source
        super(AnalyzeThread, self).__init__(parent)

    def run(self):
        t = threading.Thread(target=mdds.multi_thread_detect.main, args=(self.source, myCalbFunc))
        t.start()
        t.join()
        # self._signal.emit()

def myCalbFunc(result):
    """对检测结果进行处理的函数，可以是界面展示的回调函数，存储数据库的回调函数等
        本函数是将检测结果保存到了数据库中
        输入：
            result: [filename, fileType, BAD/GOOD, describe]
    """
    global max_num
    DBLock.acquire()
    # # print(mdds.detect.sum)
    # print(max_num)
    progress = (mdds.detect.sum / max_num) * 100
    progress_text = str(mdds.detect.sum) + "/" + str(max_num)
    #print(progress_text)
    # QMetaObject.invokeMethod(mywindow,"updateLabel",Qt.QueuedConnection, Q_ARG("int", mdds.detect.sum))
    # UIManager.ui.progressLable.setText(progress_text)
    UIManager.ui.progressBar.setValue(progress)
    UIManager.ui.progressLable.setText(progress_text)
    UIManager.ui.tableView.setModel(UIManager.model)
    result[2] = str(result[2])
    UIManager.showRecord(result)
    DBLock.release()




if __name__ == '__main__':
    app = QtGui.QApplication(sys.argv)
    UIManager = MyWindow()
    UIManager2 = MyWindow2()
    UIManager.show()
    sys.exit(app.exec_())
